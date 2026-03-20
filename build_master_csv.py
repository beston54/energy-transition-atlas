#!/usr/bin/env python3.11
"""
Build master CSV of best practices for the Energy Transition Atlas.
Sources: Excel spreadsheets + scraped WordPress pages.
"""

import csv
import re
import time
import urllib.parse
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup

# ── Paths ──
EXCEL_PATH = (
    "/Users/intern/Library/CloudStorage/OneDrive-SharedLibraries-RenewablesGridInitiative/"
    "RGI - Documents/WP3_Social Dimension/2_Communications/Website/"
    "Energy Transition Atlas/ETA_New DB_filled_Final.xlsx"
)
OUTPUT_PATH = "/Users/intern/Desktop/ETA/practices_master.csv"

# ── Sitemap URLs (extracted earlier) ──
SITEMAP_URL = "https://renewables-grid.eu/database-sitemap.xml"

# ── Theme mapping ──
THEME_MAP = {
    "Energy & Nature": "Nature",
    "Energy & Society": "People",
    "Grids & Energy Systems": "Technology",
}

# ── CSV columns ──
COLUMNS = [
    "id", "title", "url", "brand", "theme", "topic", "inf", "year",
    "country", "org", "desc", "highlights", "about",
    "img", "img2", "img3",
    "button_url", "button_url2", "button_url3",
    "video", "award", "source",
]


def map_theme(raw_dim):
    """Map old dimension names to new theme names, handling semicolons."""
    if not raw_dim:
        return ""
    parts = [p.strip() for p in raw_dim.split(";")]
    mapped = []
    for p in parts:
        m = THEME_MAP.get(p, p)
        if m not in mapped:
            mapped.append(m)
    return ", ".join(mapped)


def clean_text(val):
    """Clean cell value to string."""
    if val is None:
        return ""
    s = str(val).strip()
    # Normalize whitespace
    s = re.sub(r'\s+', ' ', s)
    return s


def clean_url(val):
    """Clean and strip URL."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith('\n'):
        s = s.strip()
    return s


def normalize_url(url):
    """Normalize URL for dedup comparison."""
    url = url.strip().rstrip("/").lower()
    url = url.replace("http://", "https://")
    # Remove query strings for comparison
    url = url.split("?")[0].split("#")[0]
    return url


def read_excel_sheet(wb, sheet_name):
    """Read practices from an Excel sheet."""
    ws = wb[sheet_name]
    practices = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        title = clean_text(row[7])
        if not title:
            continue

        url = clean_url(row[0])
        brand = clean_text(row[1])
        inf = clean_text(row[2])
        raw_dim = clean_text(row[3])
        raw_topic = clean_text(row[4])
        # row[5] = Tags, row[6] = Webinar/Video
        video = clean_url(row[6])
        year_val = row[8]
        org = clean_text(row[9])
        desc = clean_text(row[10])
        highlights = clean_text(row[11])
        about = clean_text(row[12])
        button_url = clean_url(row[13])
        button_url2 = clean_url(row[14])
        button_url3 = clean_url(row[15])
        img = clean_url(row[16])
        img2 = clean_url(row[17])
        img3 = clean_url(row[18])
        # row[19] = Comments
        award_raw = clean_text(row[20])

        # Map theme
        theme = map_theme(raw_dim)

        # Normalize topic (handle semicolons -> commas)
        topic = raw_topic.replace(";", ",") if raw_topic else ""
        # Clean up spacing
        topic = ", ".join(t.strip() for t in topic.split(",") if t.strip())

        # Parse year
        year = ""
        if year_val is not None:
            try:
                year = int(year_val)
            except (ValueError, TypeError):
                year = str(year_val).strip()

        # Parse award
        award = "true" if award_raw and "award" in award_raw.lower() else "false"

        practices.append({
            "title": title,
            "url": url,
            "brand": brand,
            "theme": theme,
            "topic": topic,
            "inf": inf,
            "year": year,
            "country": "",  # Not in Excel — will try to extract from website
            "org": org,
            "desc": desc,
            "highlights": highlights,
            "about": about,
            "img": img,
            "img2": img2,
            "img3": img3,
            "button_url": button_url,
            "button_url2": button_url2,
            "button_url3": button_url3,
            "video": video,
            "award": award,
        })
    return practices


def get_sitemap_urls():
    """Fetch all practice URLs from the WordPress sitemap."""
    print("Fetching sitemap...")
    resp = requests.get(SITEMAP_URL, timeout=30)
    soup = BeautifulSoup(resp.text, "html.parser")
    urls = []
    for loc in soup.find_all("loc"):
        url = loc.text.strip()
        # Skip the main /database/ page
        if url.rstrip("/").endswith("/database"):
            continue
        urls.append(url)
    print(f"  Found {len(urls)} practice URLs in sitemap")
    return urls


def scrape_practice(url, session):
    """Scrape a single practice page for metadata."""
    try:
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"  ERROR fetching {url}: {e}")
        return None

    soup = BeautifulSoup(resp.text, "html.parser")

    # Title — use separator to avoid joined words from line breaks
    title_el = soup.find("h1")
    title = title_el.get_text(separator=" ", strip=True) if title_el else ""
    # Collapse multiple spaces
    title = re.sub(r'\s+', ' ', title).strip()

    country = ""
    year = ""
    org = ""
    desc = ""
    img = ""
    topic = ""
    theme = ""

    # ── RGI theme structure: single-pages-meta__item with icon imgs ──
    for item in soup.find_all("div", class_="single-pages-meta__item"):
        icon = item.find("img")
        if not icon:
            continue
        alt = (icon.get("alt") or "").lower()
        text = item.get_text(strip=True)
        if "location" in alt and not country:
            country = text
        elif "calendar" in alt and not year:
            m = re.search(r'(\d{4})', text)
            if m:
                year = m.group(1)

    # ── Organisation from single-pages-meta__text ──
    org_el = soup.find("p", class_="single-pages-meta__text")
    if org_el:
        org_text = org_el.get_text(strip=True)
        # Format: "Organisation: NAME"
        m = re.match(r'Organisation\s*:\s*(.+)', org_text)
        if m:
            org = m.group(1).strip()

    # ── Topic from /topics/ links (first one on the page in the main content) ──
    topic_links = []
    for a in soup.find_all("a", href=re.compile(r'/topics/')):
        t = a.get_text(strip=True)
        href = a.get("href", "")
        if t and t != "Topics" and t not in topic_links:
            topic_links.append(t)
    # Use the first unique topic link that appears in the main practice area
    if topic_links:
        topic = topic_links[0]

    # ── Map scraped topic to theme ──
    TOPIC_TO_THEME = {
        "Bird Protection": "Nature",
        "Integrated Vegetation Management": "Nature",
        "Nature Conservation & Restoration": "Nature",
        "Offshore Energy & Nature": "Nature",
        "Monitoring & Reporting": "Nature",
        "Circularity & Supply Chains": "Technology",
        "Circularity and Supply Chains": "Technology",
        "Climate Adaptation & Resilience": "Technology",
        "Energy System Planning": "Technology",
        "Energy System Planning & Optimisation": "Technology",
        "Spatial Optimisation": "Technology",
        "Public Acceptance & Engagement": "People",
        "Creating awareness & capacity-building": "People",
        "Creating Awareness & Capacity Building": "People",
        "Fair & Inclusive Energy Transition": "People",
        "Advocating for Optimised Grids": "Planning",
        "Implementing RGI Declarations": "Planning",
        "Energy System Optimisation": "Technology",
        "Energy System Planning & Optimisation": "Technology",
    }
    if topic:
        # Try exact match first, then case-insensitive
        theme = TOPIC_TO_THEME.get(topic, "")
        if not theme:
            for key, val in TOPIC_TO_THEME.items():
                if key.lower() == topic.lower():
                    theme = val
                    break

    # ── Featured image: prefer gallery slider photos (actual practice images) ──
    gallery_pic = soup.find("picture", class_="gallery-slider__picture")
    if gallery_pic:
        gallery_img = gallery_pic.find("img")
        if gallery_img and gallery_img.get("src"):
            img = gallery_img["src"]
    # Fallback to og:image only if no gallery image found
    if not img:
        og_img = soup.find("meta", property="og:image")
        if og_img and og_img.get("content"):
            img = og_img["content"]

    # ── Description from og:description or first content paragraph ──
    og_desc = soup.find("meta", property="og:description")
    if og_desc and og_desc.get("content"):
        desc = og_desc["content"]

    if not desc:
        # Look for main content paragraphs
        for p in soup.find_all("p"):
            text = p.get_text(strip=True)
            if len(text) > 50 and "cookie" not in text.lower():
                desc = text
                break

    # Clean up desc
    if desc:
        desc = re.sub(r'\s+', ' ', desc).strip()
        if len(desc) > 500:
            desc = desc[:497] + "..."

    return {
        "title": title,
        "url": url,
        "brand": "RGI",
        "theme": theme,
        "topic": topic,
        "inf": "",
        "year": year,
        "country": country,
        "org": org,
        "desc": desc,
        "highlights": "",
        "about": "",
        "img": img,
        "img2": "",
        "img3": "",
        "button_url": "",
        "button_url2": "",
        "button_url3": "",
        "video": "",
        "award": "false",
    }


def main():
    # ── Step 1: Read Excel data ──
    print("Reading Excel data...")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    new_website = read_excel_sheet(wb, "New Website")
    print(f"  New Website: {len(new_website)} practices")

    old_website = read_excel_sheet(wb, "Old Website(not accessible)")
    print(f"  Old Website: {len(old_website)} practices")

    # Deduplicate: prefer New Website version
    by_title = {}
    # Add old first so new overwrites
    for p in old_website:
        key = p["title"].lower().strip()
        p["source"] = "old_website"
        by_title[key] = p
    for p in new_website:
        key = p["title"].lower().strip()
        p["source"] = "new_website"
        by_title[key] = p

    # Build title index for dedup with scraped data
    title_index = set(by_title.keys())

    print(f"  Combined unique (by title): {len(by_title)} practices")

    # ── Step 2: Scrape ALL website practices for metadata ──
    sitemap_urls = get_sitemap_urls()

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) ETA-DataCollector/1.0"
    })

    new_count = 0
    enriched_count = 0
    for i, url in enumerate(sitemap_urls):
        slug = url.rstrip("/").split("/")[-1]
        print(f"  Scraping [{i+1}/{len(sitemap_urls)}]: {slug}")
        result = scrape_practice(url, session)
        if not result:
            continue

        key = result["title"].lower().strip()
        if not key:
            continue

        if key in by_title:
            # Enrich existing practice with scraped data where missing
            existing = by_title[key]
            if not existing["country"] and result["country"]:
                existing["country"] = result["country"]
                enriched_count += 1
            if not existing["year"] and result["year"]:
                existing["year"] = result["year"]
            if not existing["org"] and result["org"]:
                existing["org"] = result["org"]
            if not existing["img"] and result["img"]:
                existing["img"] = result["img"]
            if not existing["topic"] and result["topic"]:
                existing["topic"] = result["topic"]
            if not existing["theme"] and result["theme"]:
                existing["theme"] = result["theme"]
            # Update URL to new website URL if old-style
            if "detail=" in existing["url"]:
                existing["url"] = url
        else:
            result["source"] = "scraped"
            by_title[key] = result
            new_count += 1

        # Be respectful
        if i < len(sitemap_urls) - 1:
            time.sleep(0.3)

    print(f"  New practices from scrape: {new_count}")
    print(f"  Existing practices enriched with country: {enriched_count}")

    all_practices = list(by_title.values())

    # Sort: year desc, then title asc
    def sort_key(p):
        try:
            y = -int(p["year"]) if p["year"] else 0
        except (ValueError, TypeError):
            y = 0
        return (y, p["title"].lower())

    all_practices.sort(key=sort_key)

    # Assign IDs
    for i, p in enumerate(all_practices, 1):
        p["id"] = i

    # ── Step 4: Write CSV ──
    print(f"\nWriting {len(all_practices)} practices to {OUTPUT_PATH}...")
    with open(OUTPUT_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_practices)

    print("Done!")

    # ── Stats ──
    sources = {}
    themes = {}
    for p in all_practices:
        s = p.get("source", "unknown")
        sources[s] = sources.get(s, 0) + 1
        for t in p.get("theme", "").split(","):
            t = t.strip()
            if t:
                themes[t] = themes.get(t, 0) + 1

    print(f"\nBy source: {sources}")
    print(f"By theme: {themes}")

    # Count with/without key fields
    has_year = sum(1 for p in all_practices if p["year"])
    has_country = sum(1 for p in all_practices if p["country"])
    has_desc = sum(1 for p in all_practices if p["desc"])
    has_img = sum(1 for p in all_practices if p["img"])
    print(f"\nField coverage:")
    print(f"  year: {has_year}/{len(all_practices)}")
    print(f"  country: {has_country}/{len(all_practices)}")
    print(f"  desc: {has_desc}/{len(all_practices)}")
    print(f"  img: {has_img}/{len(all_practices)}")


if __name__ == "__main__":
    main()
